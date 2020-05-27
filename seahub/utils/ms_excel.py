# Copyright (c) 2012-2016 Seafile Ltd.
import logging
import datetime

import openpyxl
from openpyxl.styles import Alignment

from seahub.dtable.constants import DTABLE_COL_TYPE_NUMBER, DTABLE_COL_TYPE_DATE, \
    DTABLE_COL_TYPE_GEOLOCATION, DTABLE_COL_TYPE_COLLABORATOR
from seahub.base.templatetags.seahub_tags import email2nickname


logger = logging.getLogger(__name__)

def cell_data2str(cell_data):
    if isinstance(cell_data, list):
        return '\n'.join(cell_data2str(item) for item in cell_data)
    else:
        return str(cell_data)

def is_int_str(num):
    return '.' not in str(num)


def gen_decimal_format(num):
    if is_int_str(num):
        return '0'

    decimal_cnt = len(str(num).split('.')[1])
    return '0.' + '0' * decimal_cnt


def write_xls(sheet_name, head, data_list):
    """write listed data into excel
    """

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
    except Exception as e:
        logger.error(e)
        return None

    ws.title = sheet_name

    row_num = 0

    # write table head
    for col_num in range(len(head)):
        c = ws.cell(row = row_num + 1, column = col_num + 1)
        c.value = head[col_num]

    # write table data
    for row in data_list:
        row_num += 1
        for col_num in range(len(row)):
            c = ws.cell(row = row_num + 1, column = col_num + 1)
            c.value = row[col_num]

    return wb


def write_xls_with_type(sheet_name, head, data_list):
    """ write listed data into excel
        head is a list of tuples,
        e.g. head = [(col_name, col_type, col_date), (...), ...]
    """

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
    except Exception as e:
        logger.error(e)
        return None

    ws.title = sheet_name

    row_num = 0

    # write table head
    for col_num in range(len(head)):
        c = ws.cell(row = row_num + 1, column = col_num + 1)
        c.value = head[col_num][0]
        # if has custem data format, set width to format len
        if isinstance(head[col_num][2], dict):
            format_str = head[col_num][2].get('format', '')
            if format_str:
                ws.column_dimensions[chr(col_num + ord('A'))].width = len(format_str) * 2
    # write table data
    for row in data_list:
        row_num += 1
        for col_num in range(len(row)):
            c = ws.cell(row = row_num + 1, column = col_num + 1)
            # excel format see
            # https://support.office.com/en-us/article/Number-format-codes-5026bbd6-04bc-48cd-bf33-80f18b4eae68
            if head[col_num][1] == DTABLE_COL_TYPE_NUMBER:
                # if value cannot convert to float or int, just pass, e.g. empty srt ''
                try:
                    if is_int_str(row[col_num]):
                        c.value = int(row[col_num])
                    else:
                        c.value = float(row[col_num])
                except Exception as e:
                    pass
                c.number_format = gen_decimal_format(c.value)
            elif head[col_num][1] == DTABLE_COL_TYPE_DATE:
                value_datetime = datetime.datetime.strptime(row[col_num], '%Y-%m-%d')
                c.value = value_datetime
                c.number_format = head[col_num][2].get('format', '')
            elif head[col_num][1] == DTABLE_COL_TYPE_GEOLOCATION:
                geo_str = ''
                for k, v in row[col_num].items():
                    geo_str += '{}: {}\n'.format(str(k), str(v))
                c.value = geo_str
                c.alignment = Alignment(wrapText=True)
            elif head[col_num][1] == DTABLE_COL_TYPE_COLLABORATOR:
                c.value = email2nickname(cell_data2str(row[col_num]))
            else:
                c.value = cell_data2str(row[col_num])

    return wb
